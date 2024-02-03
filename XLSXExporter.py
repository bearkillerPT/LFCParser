import openpyxl

class XLSXExporter:
    def __init__(self, filename):
        ## initializing the xlsx
        self.filename = filename
        self.xlsx = openpyxl.Workbook()
        self.tabCounter = 0
        
    def export(self, lfc_id, data):
        # Create a new sheet with the lfc_id
        if self.tabCounter == 0:
            self.tabCounter += 1
            sheet = self.xlsx.active
            sheet.title = lfc_id
        else:
            sheet = self.xlsx.create_sheet(lfc_id)
        
        for i, line in enumerate(data.split('\n')):
            for j, value in enumerate(line.split(';')):
                sheet.cell(row=i+1, column=j+1, value=value)
    def save(self):
        self.xlsx.save(self.filename)